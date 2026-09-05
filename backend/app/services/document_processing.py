# ruff: noqa: E501
from __future__ import annotations

import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.elongations.theory import parse_theory_candidates


def _run(args: list[str], timeout: int = 90) -> str:
    try:
        result = subprocess.run(
            args,
            check=True,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except FileNotFoundError as exc:
        raise RuntimeError("El motor de lectura de documentos no está instalado") from exc
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError("El documento tardó demasiado en procesarse") from exc
    except subprocess.CalledProcessError as exc:
        detail = (exc.stderr or "").strip()[:300]
        raise RuntimeError(detail or "No fue posible leer el documento") from exc
    return result.stdout


def extract_text(path: Path, mime_type: str, max_pdf_pages: int = 25) -> str:
    if mime_type == "application/pdf":
        text = _run(["pdftotext", "-layout", str(path), "-"])
        if len(text.strip()) >= 30:
            return text[:60000]
        with tempfile.TemporaryDirectory(prefix="obrixapy-ocr-") as tmp:
            prefix = Path(tmp) / "page"
            _run(
                [
                    "pdftoppm",
                    "-f",
                    "1",
                    "-l",
                    str(max_pdf_pages),
                    "-png",
                    "-r",
                    "180",
                    str(path),
                    str(prefix),
                ],
                timeout=180,
            )
            pages = sorted(Path(tmp).glob("page-*.png"))
            return "\n".join(
                _run(["tesseract", str(page), "stdout", "-l", "spa+eng"]) for page in pages
            )[:60000]
    return _run(["tesseract", str(path), "stdout", "-l", "spa+eng"])[:60000]


def parse_elongation_rows(text: str) -> list[dict[str, object]]:
    """Legacy adapter using the V2 semantic parser instead of positional numbers."""

    return [
        {
            "label": candidate.label,
            "classification": "unknown",
            "length_m": candidate.length_m,
            "strand_count": candidate.strand_count,
            "calculated_elongation": candidate.calculated_elongation_cm,
            "confidence": candidate.confidence,
            "source_location_json": candidate.source_location(),
        }
        for candidate in parse_theory_candidates(text)
    ]


def build_xlsx(rows: list[dict[str, object]]) -> bytes:
    """Legacy export adapter; formulas stay formulas instead of fixed tolerance values."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Elongaciones"
    headers = [
        "Etiqueta",
        "Clasificación",
        "Longitud (m)",
        "Cordones",
        "Elongación calculada (cm)",
        "Max. (cm)",
        "Elong. Medida (cm)",
        "Min. (cm)",
        "Revisión",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EDE9FE")
    for row_number, row in enumerate(rows, start=2):
        worksheet.append(
            [
                row.get("label"),
                row.get("classification"),
                row.get("length_m"),
                row.get("strand_count"),
                row.get("calculated_elongation"),
                f"=E{row_number}+(E{row_number}*0.07)",
                row.get("measured_elongation"),
                f"=E{row_number}-(E{row_number}*0.07)",
                row.get("review_status"),
            ]
        )
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 16
    for column in ("C", "E", "F", "G", "H"):
        worksheet.column_dimensions[column].width = 22
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
