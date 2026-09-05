"""Validación, lectura y reconstrucción segura de la plantilla XLSX V2."""

from __future__ import annotations

import re
import unicodedata
from copy import copy
from dataclasses import asdict, dataclass
from decimal import Decimal
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

MAX_TEMPLATE_BYTES = 20 * 1024 * 1024
MAX_TEMPLATE_UNCOMPRESSED_BYTES = 150 * 1024 * 1024
MAX_TEMPLATE_COMPRESSION_RATIO = 120
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FORMULA_REFERENCE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")


class TemplateValidationError(ValueError):
    """The uploaded workbook is unsafe or cannot supply the required operational structure."""


@dataclass(frozen=True)
class TemplateSection:
    name: str
    section_row: int
    header_row: int
    body_start_row: int
    body_end_row: int
    formula_seed_row: int


@dataclass(frozen=True)
class TemplateMapping:
    sheet_name: str
    sections: dict[str, TemplateSection]
    columns: dict[str, int]
    formula_seeds: dict[str, str]
    tolerance_percent: Decimal
    warnings: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "sections": {name: asdict(section) for name, section in self.sections.items()},
            "columns": self.columns,
            "formula_seeds": self.formula_seeds,
            "tolerance_percent": str(self.tolerance_percent),
            "warnings": list(self.warnings),
        }


def _normalise(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Za-z0-9]+", " ", text.upper())
    return " ".join(text.split())


def validate_template_bytes(content: bytes, filename: str | None = None) -> None:
    """Reject malformed OOXML, macros, external links and compressed bombs before parsing."""

    if not content:
        raise TemplateValidationError("La plantilla XLSX está vacía")
    if len(content) > MAX_TEMPLATE_BYTES:
        raise TemplateValidationError("La plantilla XLSX supera el tamaño máximo permitido")
    if filename and not filename.lower().endswith(".xlsx"):
        raise TemplateValidationError("La plantilla debe tener extensión .xlsx")
    try:
        with ZipFile(BytesIO(content)) as archive:
            infos = archive.infolist()
            if not infos:
                raise TemplateValidationError("La plantilla XLSX no contiene archivos")
            total_uncompressed = sum(info.file_size for info in infos)
            if total_uncompressed > MAX_TEMPLATE_UNCOMPRESSED_BYTES:
                raise TemplateValidationError("La plantilla XLSX se expande demasiado al abrirse")
            if total_uncompressed > len(content) * MAX_TEMPLATE_COMPRESSION_RATIO:
                raise TemplateValidationError("La compresión de la plantilla XLSX no es segura")
            names = set(archive.namelist())
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                raise TemplateValidationError("El archivo no es una plantilla XLSX válida")
            for name in names:
                path = PurePosixPath(name)
                if path.is_absolute() or ".." in path.parts:
                    raise TemplateValidationError("La plantilla XLSX contiene rutas inválidas")
                if name.lower().endswith("vbaProject.bin".lower()):
                    raise TemplateValidationError("No se permiten macros en la plantilla")
            for name in names:
                if name.endswith(".rels") or "externalLinks" in name:
                    relation = archive.read(name)
                    if b'TargetMode="External"' in relation or b"externalLink" in relation:
                        raise TemplateValidationError(
                            "No se permiten enlaces externos en la plantilla"
                        )
    except BadZipFile as exc:
        raise TemplateValidationError("El archivo no es un XLSX ZIP válido") from exc


def _cell_containing(ws: Any, predicate: Any) -> tuple[int, int] | None:
    for row in ws.iter_rows():
        for cell in row:
            if predicate(_normalise(cell.value)):
                return cell.row, cell.column
    return None


def _find_section(ws: Any, expected: str) -> int:
    match = _cell_containing(ws, lambda value: value == expected)
    if match is None:
        raise TemplateValidationError(
            f"No se encontró la sección {expected.title()} en la plantilla"
        )
    return match[0]


def _find_header_row(ws: Any, section_row: int) -> int:
    for row_number in range(section_row + 1, min(ws.max_row, section_row + 10) + 1):
        values = [
            _normalise(ws.cell(row_number, column).value)
            for column in range(1, ws.max_column + 1)
        ]
        joined = " ".join(values)
        if "ITEM" in joined and "LABEL" in joined and "LONGITUD" in joined and "CANTIDAD" in joined:
            return row_number
    raise TemplateValidationError("No se localizaron las columnas obligatorias de la plantilla")


def _find_columns(ws: Any, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for row_number in (header_row, header_row + 1):
        for column in range(1, ws.max_column + 1):
            value = _normalise(ws.cell(row_number, column).value)
            if value == "ITEM":
                columns.setdefault("item", column)
            elif value == "LABEL":
                columns.setdefault("label", column)
            elif "LONGITUD" in value:
                columns.setdefault("length_m", column)
            elif "CANTIDAD" in value and "TENDON" in value:
                columns.setdefault("strand_count", column)
            elif value == "CALCULADA":
                columns.setdefault("calculated", column)
            elif value.startswith("MAX"):
                columns.setdefault("maximum", column)
            elif "MEDIDA" in value:
                columns.setdefault("measured", column)
            elif value.startswith("MIN"):
                columns.setdefault("minimum", column)
    missing = {
        "item",
        "label",
        "length_m",
        "strand_count",
        "calculated",
        "maximum",
        "measured",
        "minimum",
    } - set(columns)
    if missing:
        raise TemplateValidationError(
            f"Faltan columnas requeridas en la plantilla: {', '.join(sorted(missing))}"
        )
    return columns


def _formula_references_own_calculated_row(formula: str, calculated_column: int, row: int) -> bool:
    if not formula.startswith("=") or "#REF!" in formula.upper():
        return False
    expected = get_column_letter(calculated_column)
    references = FORMULA_REFERENCE.findall(formula.upper())
    return bool(references) and all(
        column == expected and int(reference_row) == row
        for column, reference_row in references
    )


def _formula_signature(formula: str, origin: str, destination: str) -> str:
    return Translator(formula, origin=origin).translate_formula(destination)


def _formula_tolerance(formula: str) -> Decimal | None:
    match = re.search(r"\*\s*\(?\s*(0[.,]\d+|\d+[.,]\d+)\s*\)?", formula)
    if not match:
        return None
    return Decimal(match.group(1).replace(",", ".")) * Decimal("100")


def _find_formula_seed(
    ws: Any,
    start_row: int,
    end_row: int,
    columns: dict[str, int],
    warnings: list[str],
) -> tuple[int, dict[str, str], Decimal]:
    formula_pairs: list[tuple[int, str, str]] = []
    broken = 0
    for row in range(start_row, end_row + 1):
        maximum = ws.cell(row, columns["maximum"]).value
        minimum = ws.cell(row, columns["minimum"]).value
        if not isinstance(maximum, str) or not isinstance(minimum, str):
            continue
        if not (
            _formula_references_own_calculated_row(maximum, columns["calculated"], row)
            and _formula_references_own_calculated_row(minimum, columns["calculated"], row)
        ):
            broken += 1
            continue
        max_signature = _formula_signature(
            maximum,
            f"{get_column_letter(columns['maximum'])}{row}",
            f"{get_column_letter(columns['maximum'])}1",
        )
        min_signature = _formula_signature(
            minimum,
            f"{get_column_letter(columns['minimum'])}{row}",
            f"{get_column_letter(columns['minimum'])}1",
        )
        formula_pairs.append((row, max_signature, min_signature))
    if broken:
        warnings.append(f"Se detectaron {broken} fórmulas Max./Min. heredadas inválidas")
    if not formula_pairs:
        raise TemplateValidationError("No existe una fila con fórmulas Max./Min. válidas")
    counts: dict[tuple[str, str], int] = {}
    for _, max_signature, min_signature in formula_pairs:
        counts[(max_signature, min_signature)] = counts.get((max_signature, min_signature), 0) + 1
    (dominant_max, dominant_min), dominant_count = max(counts.items(), key=lambda entry: entry[1])
    if dominant_count * 2 <= len(formula_pairs):
        raise TemplateValidationError("No existe una regla Max./Min. dominante y segura")
    seed_row = next(
        row
        for row, max_signature, min_signature in formula_pairs
        if (max_signature, min_signature) == (dominant_max, dominant_min)
    )
    tolerance = _formula_tolerance(ws.cell(seed_row, columns["maximum"]).value)
    if tolerance is None or tolerance < 0 or tolerance > 100:
        raise TemplateValidationError("No fue posible determinar la tolerancia desde la fórmula")
    return (
        seed_row,
        {
            "maximum": dominant_max,
            "minimum": dominant_min,
        },
        tolerance,
    )


def analyse_template(content: bytes, filename: str | None = None) -> TemplateMapping:
    """Discover visual structure and a verified dominant formula, never fixed cell addresses."""

    validate_template_bytes(content, filename)
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_vba=False)
    except Exception as exc:  # openpyxl groups multiple malformed OOXML exceptions
        raise TemplateValidationError("No fue posible abrir la plantilla XLSX") from exc
    worksheet = workbook.active
    band_row = _find_section(worksheet, "BANDAS")
    distributed_row = _find_section(worksheet, "DISTRIBUIDOS")
    if distributed_row <= band_row:
        raise TemplateValidationError(
            "Las secciones BANDAS y DISTRIBUIDOS están en un orden inválido"
        )
    band_header = _find_header_row(worksheet, band_row)
    distributed_header = _find_header_row(worksheet, distributed_row)
    columns = _find_columns(worksheet, band_header)
    distributed_columns = _find_columns(worksheet, distributed_header)
    if columns != distributed_columns:
        raise TemplateValidationError(
            "Las secciones de la plantilla no comparten las mismas columnas"
        )
    warnings: list[str] = []
    band_start = band_header + 2
    distributed_start = distributed_header + 2
    band_seed, formulas, tolerance = _find_formula_seed(
        worksheet, band_start, distributed_row - 1, columns, warnings
    )
    distributed_seed, distributed_formulas, distributed_tolerance = _find_formula_seed(
        worksheet, distributed_start, worksheet.max_row, columns, warnings
    )
    if tolerance != distributed_tolerance or formulas != distributed_formulas:
        warnings.append(
            "La sección Distribuidos usa una fórmula diferente; se aplicará la regla dominante"
        )
    return TemplateMapping(
        sheet_name=worksheet.title,
        sections={
            "band": TemplateSection(
                "band", band_row, band_header, band_start, distributed_row - 1, band_seed
            ),
            "distributed": TemplateSection(
                "distributed",
                distributed_row,
                distributed_header,
                distributed_start,
                worksheet.max_row,
                distributed_seed,
            ),
        },
        columns=columns,
        formula_seeds=formulas,
        tolerance_percent=tolerance.quantize(Decimal("0.01")),
        warnings=tuple(warnings),
    )


def _prototype(ws: Any, row: int, columns: dict[str, int]) -> dict[str, Any]:
    return {
        "cells": {
            column: copy(ws.cell(row, column)._style)
            for column in range(1, ws.max_column + 1)
        },
        "height": ws.row_dimensions[row].height,
        "number_formats": {
            key: ws.cell(row, column).number_format for key, column in columns.items()
        },
    }


def _unmerge_body_ranges(ws: Any, start_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row:
            ws.unmerge_cells(str(merged_range))


def _apply_prototype(ws: Any, row: int, prototype: dict[str, Any]) -> None:
    for column, style in prototype["cells"].items():
        ws.cell(row, column)._style = copy(style)
        ws.cell(row, column).value = None
    ws.row_dimensions[row].height = prototype["height"]


def _physical_rows(groups: list[dict[str, Any]]) -> list[tuple[dict[str, Any], int]]:
    rows: list[tuple[dict[str, Any], int]] = []
    for group in groups:
        count = int(group["strand_count"])
        if count <= 0:
            raise TemplateValidationError(f"{group['label']} tiene una cantidad S inválida")
        rows.extend((group, ordinal) for ordinal in range(1, count + 1))
    return rows


def _formula_for_row(seed: str, source_column: int, target_column: int, row: int) -> str:
    translated = Translator(seed, origin=f"{get_column_letter(target_column)}1").translate_formula(
        f"{get_column_letter(target_column)}{row}"
    )
    if not _formula_references_own_calculated_row(translated, source_column, row):
        raise TemplateValidationError(
            "La fórmula trasladada no apunta a Calculada de su propia fila"
        )
    return translated


def _write_section(
    ws: Any,
    mapping: TemplateMapping,
    section_name: str,
    start_row: int,
    groups: list[dict[str, Any]],
    prototype: dict[str, Any],
    *,
    final: bool,
) -> int:
    columns = mapping.columns
    row_number = start_row
    item_number = 1
    for group, ordinal in _physical_rows(groups):
        _apply_prototype(ws, row_number, prototype)
        measurements = {int(value["ordinal"]): value for value in group.get("measurements", [])}
        measurement = measurements.get(ordinal)
        ws.cell(row_number, columns["item"]).value = item_number
        ws.cell(row_number, columns["calculated"]).value = group["calculated_elongation"]
        ws.cell(row_number, columns["maximum"]).value = _formula_for_row(
            mapping.formula_seeds["maximum"],
            columns["calculated"],
            columns["maximum"],
            row_number,
        )
        ws.cell(row_number, columns["minimum"]).value = _formula_for_row(
            mapping.formula_seeds["minimum"],
            columns["calculated"],
            columns["minimum"],
            row_number,
        )
        ws.cell(row_number, columns["measured"]).value = (
            measurement.get("measured_elongation") if final and measurement else None
        )
        for key, column in prototype["number_formats"].items():
            ws.cell(row_number, columns[key]).number_format = column
        row_number += 1
        item_number += 1
    row_number = start_row
    for group in groups:
        physical_count = int(group["strand_count"])
        ws.cell(row_number, columns["label"]).value = group["label"]
        ws.cell(row_number, columns["length_m"]).value = group["length_m"]
        ws.cell(row_number, columns["strand_count"]).value = physical_count
        if physical_count > 1:
            end_row = row_number + physical_count - 1
            for key in ("label", "length_m", "strand_count"):
                column = get_column_letter(columns[key])
                ws.merge_cells(f"{column}{row_number}:{column}{end_row}")
        row_number += physical_count
    return row_number


def _assert_export_groups(groups: list[dict[str, Any]], final: bool) -> None:
    labels: set[str] = set()
    for group in groups:
        label = str(group["label"])
        if label in labels:
            raise TemplateValidationError(f"La etiqueta {label} está duplicada")
        labels.add(label)
        if group.get("classification") not in {"band", "distributed"}:
            raise TemplateValidationError(f"La etiqueta {label} sigue sin clasificar")
        measurements = group.get("measurements", [])
        if final:
            expected = int(group["strand_count"])
            if len(measurements) != expected:
                raise TemplateValidationError(f"{label} no tiene exactamente S mediciones")
            for measurement in measurements:
                if measurement.get("measured_elongation") is None:
                    raise TemplateValidationError(f"{label} tiene mediciones faltantes")
                if measurement.get("review_status") != "approved":
                    raise TemplateValidationError(
                        f"{label} tiene mediciones sin aprobación técnica"
                    )
                if (
                    measurement.get("tolerance_status") == "outside"
                    and not measurement.get("override_reason")
                ):
                    raise TemplateValidationError(
                        f"{label} tiene una excepción fuera de tolerancia sin observación"
                    )


def _add_control_sheets(
    workbook: Workbook,
    groups: list[dict[str, Any]],
    *,
    history: dict[str, Any],
) -> None:
    for name in ("Control OCR", "Historial Obrixapy"):
        if name in workbook.sheetnames:
            del workbook[name]
    control = workbook.create_sheet("Control OCR")
    control.append(
        [
            "Label",
            "Ordinal",
            "Valor (cm)",
            "Archivo/página",
            "Recorte",
            "Confianza",
            "Asociación",
            "Revisión",
            "Tolerancia",
            "Observación",
        ]
    )
    for group in sorted(groups, key=lambda item: int(item.get("label_number", 0))):
        for measurement in sorted(group.get("measurements", []), key=lambda item: item["ordinal"]):
            location = measurement.get("source_location_json") or {}
            control.append(
                [
                    group["label"],
                    measurement["ordinal"],
                    measurement.get("measured_elongation"),
                    location.get("file") or location.get("page"),
                    location.get("bbox"),
                    measurement.get("confidence"),
                    measurement.get("match_method"),
                    measurement.get("review_status"),
                    measurement.get("tolerance_status"),
                    measurement.get("override_reason"),
                ]
            )
    history_sheet = workbook.create_sheet("Historial Obrixapy")
    history_sheet.append(
        ["Trabajo", "Versión", "Tipo", "Fecha", "Usuario", "Fuentes", "SHA salida"]
    )
    history_sheet.append(
        [
            history.get("job_title"),
            history.get("version_number"),
            history.get("kind"),
            history.get("created_at"),
            history.get("created_by"),
            history.get("source_hashes"),
            history.get("output_sha256"),
        ]
    )
    for sheet in (control, history_sheet):
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = copy(workbook.active[1][0].font)
            cell.fill = PatternFill("solid", fgColor="EDE9FE")
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 20


def build_export_xlsx(
    template_content: bytes,
    mapping: TemplateMapping,
    groups: list[dict[str, Any]],
    *,
    final: bool,
    history: dict[str, Any],
) -> bytes:
    """Rebuild the operational body while keeping a template's header and presentation.

    Every physical tendon receives its own row.  Max./Min. stay as translated Excel formulas in
    every row and are validated again before returning the immutable export bytes.
    """

    _assert_export_groups(groups, final)
    validate_template_bytes(template_content)
    workbook = load_workbook(BytesIO(template_content), data_only=False, keep_vba=False)
    ws = workbook[mapping.sheet_name]
    band_groups = sorted(
        [group for group in groups if group["classification"] == "band"],
        key=lambda item: int(item.get("label_number", 0)),
    )
    distributed_groups = sorted(
        [group for group in groups if group["classification"] == "distributed"],
        key=lambda item: int(item.get("label_number", 0)),
    )
    band_section = mapping.sections["band"]
    distributed_section = mapping.sections["distributed"]
    band_prototype = _prototype(ws, band_section.formula_seed_row, mapping.columns)
    distributed_prototype = _prototype(ws, distributed_section.formula_seed_row, mapping.columns)
    _unmerge_body_ranges(ws, band_section.body_start_row)

    band_capacity = distributed_section.section_row - band_section.body_start_row
    ws.delete_rows(band_section.body_start_row, band_capacity)
    band_rows = _physical_rows(band_groups)
    if band_rows:
        ws.insert_rows(band_section.body_start_row, len(band_rows))
    after_band_row = _write_section(
        ws,
        mapping,
        "band",
        band_section.body_start_row,
        band_groups,
        band_prototype,
        final=final,
    )

    new_distributed_section_row = after_band_row
    distributed_body_offset = distributed_section.body_start_row - distributed_section.section_row
    new_distributed_body_row = new_distributed_section_row + distributed_body_offset
    if new_distributed_body_row <= ws.max_row:
        ws.delete_rows(new_distributed_body_row, ws.max_row - new_distributed_body_row + 1)
    distributed_rows = _physical_rows(distributed_groups)
    if distributed_rows:
        ws.insert_rows(new_distributed_body_row, len(distributed_rows))
    after_distributed_row = _write_section(
        ws,
        mapping,
        "distributed",
        new_distributed_body_row,
        distributed_groups,
        distributed_prototype,
        final=final,
    )

    for column in ("E", "F", "G", "L", "M", "N"):
        ws.column_dimensions[column].hidden = True
    measured_column = get_column_letter(mapping.columns["measured"])
    maximum_column = get_column_letter(mapping.columns["maximum"])
    minimum_column = get_column_letter(mapping.columns["minimum"])
    first_data_row = band_section.body_start_row
    last_data_row = max(first_data_row, after_distributed_row - 1)
    measured_range = f"{measured_column}{first_data_row}:{measured_column}{last_data_row}"
    ws.conditional_formatting.add(
        measured_range,
        FormulaRule(
            formula=[
                f'AND({measured_column}{first_data_row}="",'
                f'{get_column_letter(mapping.columns["calculated"])}{first_data_row}<>"")'
            ],
            fill=PatternFill("solid", fgColor="FEF3C7"),
        ),
    )
    ws.conditional_formatting.add(
        measured_range,
        FormulaRule(
            formula=[
                f'AND({measured_column}{first_data_row}<>"",'
                f'OR({measured_column}{first_data_row}<{minimum_column}{first_data_row},'
                f'{measured_column}{first_data_row}>{maximum_column}{first_data_row}))'
            ],
            fill=PatternFill("solid", fgColor="FECACA"),
        ),
    )
    ws.print_area = f"A1:K{last_data_row}"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    _add_control_sheets(workbook, groups, history=history)

    output = BytesIO()
    workbook.save(output)
    result = output.getvalue()
    _validate_generated_export(result, mapping, first_data_row, last_data_row)
    return result


def _validate_generated_export(
    content: bytes,
    mapping: TemplateMapping,
    first_data_row: int,
    last_data_row: int,
) -> None:
    """Reopen and prove formula integrity before bytes are stored as an export version."""

    workbook = load_workbook(BytesIO(content), data_only=False, keep_vba=False)
    ws = workbook[mapping.sheet_name]
    for row in range(first_data_row, last_data_row + 1):
        calculated = ws.cell(row, mapping.columns["calculated"]).value
        if calculated is None or isinstance(calculated, str):
            continue
        for key in ("maximum", "minimum"):
            formula = ws.cell(row, mapping.columns[key]).value
            if not isinstance(formula, str) or not _formula_references_own_calculated_row(
                formula, mapping.columns["calculated"], row
            ):
                raise TemplateValidationError(
                    "El Excel generado contiene una fórmula Max./Min. inválida"
                )
            if "#REF!" in formula.upper():
                raise TemplateValidationError("El Excel generado contiene una referencia rota")
