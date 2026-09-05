from __future__ import annotations

import asyncio
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from fastapi import BackgroundTasks, HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw
from starlette.datastructures import Headers

from app.api.routes import elongations
from app.api.routes.elongations import _resume_queued_theory_job, require_job
from app.db.models import ElongationJob
from app.services.elongations import theory
from app.services.elongations.classification import propose_classifications, zone_contains
from app.services.elongations.measurements import (
    _perspective_correct,
    expand_measurement_slots,
    parse_handwritten_values,
    tolerance_status,
)
from app.services.elongations.pipeline import (
    DOCUMENT_APPROVER_ROLES,
    json_safe,
    progress_for,
)
from app.services.elongations.template import (
    TemplateValidationError,
    analyse_template,
    build_export_xlsx,
    validate_template_bytes,
)
from app.services.elongations.theory import (
    _map_rotated_tile_box,
    _parse_tsv_blocks,
    _pdf_layout,
    _semantic_ocr_blocks,
    _tile_offsets,
    deduplicate_candidates,
    parse_theory_candidates,
)
from app.services.file_storage import XLSX_MIME_TYPES, store_upload


def synthetic_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Operativa"
    sections = ((1, 3, 5, "BANDAS"), (15, 17, 19, "DISTRIBUIDOS"))
    for section_row, header_row, body_row, title in sections:
        worksheet.cell(section_row, 1).value = title
        worksheet.cell(header_row, 1).value = "Item"
        worksheet.cell(header_row, 2).value = "Label"
        worksheet.cell(header_row, 3).value = "Longitud (m)"
        worksheet.cell(header_row, 4).value = "Cantidad de tendones"
        worksheet.cell(header_row, 5).value = "Elongación (cm)"
        worksheet.cell(header_row + 1, 5).value = "Calculada"
        worksheet.cell(header_row + 1, 6).value = "Max."
        worksheet.cell(header_row + 1, 7).value = "Elong. Medida"
        worksheet.cell(header_row + 1, 8).value = "Min."
        worksheet.cell(body_row, 1).value = 1
        worksheet.cell(body_row, 2).value = "T1"
        worksheet.cell(body_row, 3).value = Decimal("10.000")
        worksheet.cell(body_row, 4).value = 1
        worksheet.cell(body_row, 5).value = Decimal("7.000")
        worksheet.cell(body_row, 6).value = f"=E{body_row}+(E{body_row}*0.07)"
        worksheet.cell(body_row, 8).value = f"=E{body_row}-(E{body_row}*0.07)"
        for column in (3, 5, 6, 7, 8):
            worksheet.cell(body_row, column).number_format = "0.000"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def shared_header_template() -> bytes:
    """Field layout: one header above BANDAS and no repeated header later."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "N1 (Ajustado)"
    worksheet.cell(1, 1).value = "Item"
    worksheet.cell(1, 2).value = "Label"
    worksheet.cell(1, 3).value = "Longitud (m)"
    worksheet.cell(1, 4).value = "Cantidad de tendones"
    worksheet.cell(1, 5).value = "Elongación (cm)"
    worksheet.cell(2, 5).value = "Calculada"
    worksheet.cell(2, 6).value = "Max."
    worksheet.cell(2, 7).value = "Elong. Medida"
    worksheet.cell(2, 8).value = "Min."
    for section_row, body_row, title, label in (
        (3, 5, "BANDAS", "T1"),
        (15, 17, "DISTRIBUIDOS", "T203"),
    ):
        worksheet.cell(section_row, 1).value = title
        worksheet.cell(body_row, 1).value = 1
        worksheet.cell(body_row, 2).value = label
        worksheet.cell(body_row, 3).value = Decimal("10.000")
        worksheet.cell(body_row, 4).value = 1
        worksheet.cell(body_row, 5).value = Decimal("7.000")
        worksheet.cell(body_row, 6).value = f"=E{body_row}+(E{body_row}*0.07)"
        worksheet.cell(body_row, 8).value = f"=E{body_row}-(E{body_row}*0.07)"
        for column in (3, 5, 6, 7, 8):
            worksheet.cell(body_row, column).number_format = "0.000"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_parser_requires_all_semantic_fields_and_normalises_decimals() -> None:
    candidates = parse_theory_candidates(
        "Tendon 8; S = 2; L 11.880; Elong = 7,9\n"
        "Tendón 203\nS=6\nL=31,496\nElongación:20.6"
    )
    assert [(candidate.label, candidate.strand_count) for candidate in candidates] == [
        ("T8", 2),
        ("T203", 6),
    ]
    assert candidates[0].length_m == Decimal("11.880")
    assert candidates[0].calculated_elongation_cm == Decimal("7.9")
    assert candidates[1].length_m == Decimal("31.496")
    assert parse_theory_candidates("Tendon 228; $=1; L17.929; Elong=12.0")[0].label == "T228"


def test_parser_never_turns_titles_dates_or_legend_numbers_into_groups() -> None:
    source = (
        "FECHA 20/02/2026 REVISION 3 2 1\n"
        "LEYENDA S= 3,5 L= Elong=\n"
        "TENDONES CONTINUOS 15 000 kg 12.7 mm\n"
        "Cable A banda 12,5 7 85,2\n"
    )
    assert parse_theory_candidates(source) == []


def test_duplicate_mosaic_reads_keep_identical_or_mark_conflict() -> None:
    same = parse_theory_candidates(
        "Tendon 8;S=2;L=11.880;Elong=7.9\nTendon 8;S=2;L=11.880;Elong=7.9"
    )
    assert len(same) == 1
    conflict = parse_theory_candidates(
        "Tendon 8;S=2;L=11.880;Elong=7.9\nTendon 8;S=2;L=12.880;Elong=7.9"
    )
    assert len(conflict) == 1
    assert conflict[0].conflict
    assert len(conflict[0].alternatives) == 2
    assert deduplicate_candidates(conflict)[0].conflict


def test_ocr_tsv_groups_semantic_fields_and_keeps_tile_coordinates() -> None:
    tsv = "\n".join(
        [
            "level\tpage_num\tblock_num\tpar_num\tline_num\tword_num\tleft\ttop\twidth\theight\tconf\ttext",
            "5\t1\t3\t4\t1\t1\t40\t30\t100\t20\t94\tTendon",
            "5\t1\t3\t4\t1\t2\t145\t30\t18\t20\t91\t8;",
            "5\t1\t3\t5\t2\t1\t40\t55\t40\t20\t93\tS=2;",
            "5\t1\t3\t6\t2\t2\t90\t55\t75\t20\t93\tL=11,880;",
            "5\t1\t3\t7\t3\t1\t40\t80\t85\t20\t92\tElong=7.9",
        ]
    )
    blocks = _parse_tsv_blocks(
        tsv,
        page_number=1,
        x_offset=512,
        y_offset=1024,
        page_width=4096,
        page_height=8192,
    )
    semantic_blocks = _semantic_ocr_blocks(blocks)
    candidates = [
        candidate
        for text, page, bbox, _confidence in semantic_blocks
        for candidate in parse_theory_candidates(text, page=page, bbox=bbox)
    ]
    assert [(candidate.label, candidate.strand_count) for candidate in candidates] == [("T8", 2)]
    _, _, bbox, confidence = semantic_blocks[-1]
    assert bbox is not None
    assert bbox["x"] == Decimal("0.134766")
    assert confidence > Decimal("0.90")


def test_large_pdf_tiles_cover_rotated_output_and_vector_coordinates(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    info = "Pages:          1\nPage size:      2384 x 3370 pts (A0)\nPage rot:        90\n"
    monkeypatch.setattr(theory, "_run", lambda *_args, **_kwargs: info)
    page_count, width, height = _pdf_layout(Path("rotated.pdf"), max_pdf_pages=25)
    assert (page_count, width, height) == (1, 10531, 7450)
    x_offsets = _tile_offsets(width)
    y_offsets = _tile_offsets(height)
    assert x_offsets[0] == y_offsets[0] == 0
    assert x_offsets[-1] + 4096 >= width
    assert y_offsets[-1] + 4096 >= height
    assert len(x_offsets) == len(set(x_offsets))
    assert len(x_offsets) * len(y_offsets) == 6


def test_rotated_ocr_boxes_return_to_the_original_tile_coordinates() -> None:
    assert _map_rotated_tile_box(
        10,
        20,
        30,
        40,
        tile_width=400,
        tile_height=300,
        rotation=90,
    ) == (340, 10, 40, 30)
    assert _map_rotated_tile_box(
        10,
        20,
        30,
        40,
        tile_width=400,
        tile_height=300,
        rotation=270,
    ) == (20, 260, 40, 30)


def test_geometry_proposals_are_conservative_and_zones_use_normalized_centers() -> None:
    candidates = [
        *parse_theory_candidates(
            "Tendon 8;S=1;L=10;Elong=7",
            bbox={
                "x": Decimal("0.05"),
                "y": Decimal("0.30"),
                "width": Decimal("0.02"),
                "height": Decimal("0.02"),
            },
        ),
        *parse_theory_candidates(
            "Tendon 9;S=1;L=10;Elong=7",
            bbox={
                "x": Decimal("0.10"),
                "y": Decimal("0.50"),
                "width": Decimal("0.02"),
                "height": Decimal("0.02"),
            },
        ),
    ]
    proposals = propose_classifications(candidates)
    assert proposals["T8"].classification == "band"
    assert proposals["T8"].confidence < Decimal("0.90")
    assert zone_contains(
        {"page": 1, "x": 0, "y": 0.2, "width": 0.2, "height": 0.4},
        candidates[0].source_location(),
    )
    assert not zone_contains(
        {"page": 2, "x": 0, "y": 0.2, "width": 0.2, "height": 0.4},
        candidates[0].source_location(),
    )


def test_handwritten_values_split_hyphen_lists_and_keep_excess_visible() -> None:
    values = parse_handwritten_values("Tendon 321; S=4; 4,8 - 4.0 / 4,5 4.5 - 5,1")
    assert values == [
        Decimal("4.8"),
        Decimal("4.0"),
        Decimal("4.5"),
        Decimal("4.5"),
        Decimal("5.1"),
    ]
    slots, extras = expand_measurement_slots(4, values)
    assert [slot.measured_elongation_cm for slot in slots] == values[:4]
    assert extras == [Decimal("5.1")]
    assert len(expand_measurement_slots(4, values[:2])[0]) == 4


def test_local_scan_perspective_correction_keeps_a_separate_derivative() -> None:
    source = Image.new("RGB", (400, 300), color="#4b5563")
    draw = ImageDraw.Draw(source)
    draw.polygon([(48, 20), (350, 36), (338, 270), (60, 260)], fill="white")
    corrected, applied = _perspective_correct(source)
    try:
        assert applied is True
        assert corrected.size[0] > 200
        assert corrected.size[1] > 180
        assert source.size == (400, 300)
    finally:
        source.close()
        corrected.close()


def test_tolerance_uses_decimal_and_outside_needs_review() -> None:
    assert tolerance_status(Decimal("10"), Decimal("10.7"), Decimal("7")) == "within"
    assert tolerance_status(Decimal("10"), Decimal("10.701"), Decimal("7")) == "outside"
    assert tolerance_status(Decimal("10"), None, Decimal("7")) == "missing"


def test_template_mapping_finds_sections_and_formula_tolerance_without_fixed_rows() -> None:
    mapping = analyse_template(synthetic_template(), "referencia.xlsx")
    assert mapping.sheet_name == "Operativa"
    assert mapping.columns["calculated"] == 5
    assert mapping.columns["maximum"] == 6
    assert mapping.tolerance_percent == Decimal("7.00")
    assert mapping.sections["band"].body_start_row == 5


def test_template_mapping_accepts_shared_headers_above_section_titles() -> None:
    mapping = analyse_template(shared_header_template(), "N1-ajustado.xlsx")

    assert mapping.sheet_name == "N1 (Ajustado)"
    assert mapping.sections["band"].header_row == 1
    assert mapping.sections["band"].body_start_row == 5
    assert mapping.sections["distributed"].header_row == 1
    assert mapping.sections["distributed"].body_start_row == 17
    assert mapping.columns["measured"] == 7
    assert "encabezados compartidos" in " ".join(mapping.warnings)


def test_dynamic_export_keeps_shared_header_sections_and_own_formulas() -> None:
    template = shared_header_template()
    mapping = analyse_template(template)
    groups = [
        {
            "label": "T8",
            "label_number": 8,
            "classification": "band",
            "length_m": Decimal("11.880"),
            "strand_count": 1,
            "calculated_elongation": Decimal("7.9"),
            "measurements": [{"ordinal": 1}],
        },
        {
            "label": "T203",
            "label_number": 203,
            "classification": "distributed",
            "length_m": Decimal("31.496"),
            "strand_count": 2,
            "calculated_elongation": Decimal("20.6"),
            "measurements": [{"ordinal": 1}, {"ordinal": 2}],
        },
    ]

    content = build_export_xlsx(
        template,
        mapping,
        groups,
        final=False,
        history={"kind": "theory"},
    )
    worksheet = load_workbook(BytesIO(content), data_only=False)["N1 (Ajustado)"]
    assert worksheet.cell(5, 2).value == "T8"
    assert worksheet.cell(6, 1).value == "DISTRIBUIDOS"
    assert worksheet.cell(8, 2).value == "T203"
    assert worksheet.cell(9, 2).value is None
    for row in (5, 8, 9):
        assert worksheet.cell(row, 6).value == f"=E{row}+(E{row}*0.07)"
        assert worksheet.cell(row, 8).value == f"=E{row}-(E{row}*0.07)"


def test_create_job_refreshes_server_timestamp_before_returning_response(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    created_at = datetime(2026, 9, 5, 16, 24, 27)
    plan = SimpleNamespace(
        id="plan-version-1",
        storage_key="plans/plan.pdf",
        original_filename="plan.pdf",
        mime_type="application/pdf",
        size_bytes=1,
        sha256="plan-sha",
    )
    template_upload = SimpleNamespace(
        storage_key="templates/template.xlsx",
        original_filename="template.xlsx",
        mime_type=next(iter(XLSX_MIME_TYPES)),
        size_bytes=1,
        sha256="template-sha",
    )

    class Mapping:
        tolerance_percent = Decimal("7.00")

        def to_dict(self) -> dict[str, str]:
            return {"sheet_name": "N1"}

    class Session:
        def __init__(self) -> None:
            self.scalar_values = [plan, None, None, None]
            self.objects: list[object] = []
            self.refreshed: list[ElongationJob] = []

        async def scalar(self, _statement: object) -> object | None:
            return self.scalar_values.pop(0)

        def add(self, object_: object) -> None:
            self.objects.append(object_)

        async def flush(self) -> None:
            for object_ in self.objects:
                if isinstance(object_, ElongationJob) and object_.id is None:
                    object_.id = "job-1"

        async def commit(self) -> None:
            return None

        async def refresh(self, job: ElongationJob) -> None:
            self.refreshed.append(job)
            job.created_at = created_at

    async def no_op(*_args: object, **_kwargs: object) -> None:
        return None

    async def fake_upload(*_args: object, **_kwargs: object) -> SimpleNamespace:
        return template_upload

    async def fake_response(_db: object, job: ElongationJob) -> str:
        assert job.created_at == created_at
        return "created"

    session = Session()
    monkeypatch.setattr(elongations, "require_project", no_op)
    monkeypatch.setattr(elongations, "require_level", no_op)
    monkeypatch.setattr(elongations, "require_assignee", no_op)
    monkeypatch.setattr(elongations, "store_upload", fake_upload)
    monkeypatch.setattr(
        elongations,
        "storage_path",
        lambda _key: SimpleNamespace(read_bytes=lambda: b"template"),
    )
    monkeypatch.setattr(elongations, "analyse_template", lambda *_args: Mapping())
    monkeypatch.setattr(
        elongations,
        "get_settings",
        lambda: SimpleNamespace(document_max_bytes=1),
    )
    monkeypatch.setattr(elongations, "response_for_job", fake_response)

    response = asyncio.run(
        elongations.create_elongation_job(
            project_id="project-1",
            background_tasks=BackgroundTasks(),
            access=SimpleNamespace(
                company_id="company-1",
                role="owner",
                user=SimpleNamespace(id="user-1"),
            ),
            db=session,
            title="Elongaciones nivel 1",
            template_file=UploadFile(file=BytesIO(), filename="template.xlsx"),
            plan_version_id="plan-version-1",
        )
    )

    assert response == "created"
    assert [job.id for job in session.refreshed] == ["job-1"]


def test_idempotent_queued_job_resumes_its_theory_task() -> None:
    background_tasks = BackgroundTasks()
    queued = SimpleNamespace(id="job-1", workflow_status="queued_theory")
    completed = SimpleNamespace(id="job-2", workflow_status="theory_review")

    _resume_queued_theory_job(background_tasks, queued)
    _resume_queued_theory_job(background_tasks, completed)

    assert len(background_tasks.tasks) == 1
    assert background_tasks.tasks[0].args == ("job-1",)


def test_dynamic_export_has_one_row_per_s_and_own_max_min_formulas() -> None:
    template = synthetic_template()
    mapping = analyse_template(template)
    groups = [
        {
            "label": "T8",
            "label_number": 8,
            "classification": "band",
            "length_m": Decimal("11.880"),
            "strand_count": 1,
            "calculated_elongation": Decimal("7.9"),
            "measurements": [
                {
                    "ordinal": 1,
                    "measured_elongation": Decimal("7.8"),
                    "review_status": "approved",
                    "tolerance_status": "within",
                }
            ],
        },
        {
            "label": "T203",
            "label_number": 203,
            "classification": "distributed",
            "length_m": Decimal("31.496"),
            "strand_count": 2,
            "calculated_elongation": Decimal("20.6"),
            "measurements": [
                {
                    "ordinal": 1,
                    "measured_elongation": Decimal("20.5"),
                    "review_status": "approved",
                    "tolerance_status": "within",
                },
                {
                    "ordinal": 2,
                    "measured_elongation": Decimal("20.7"),
                    "review_status": "approved",
                    "tolerance_status": "within",
                },
            ],
        },
    ]
    content = build_export_xlsx(
        template,
        mapping,
        groups,
        final=True,
        history={"job_title": "Prueba", "version_number": 1, "kind": "final"},
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    worksheet = workbook["Operativa"]
    rows = [
        row
        for row in range(1, worksheet.max_row + 1)
        if isinstance(
            worksheet.cell(row, mapping.columns["calculated"]).value,
            int | float | Decimal,
        )
    ]
    assert len(rows) == 3
    for row in rows:
        assert worksheet.cell(row, mapping.columns["maximum"]).value == f"=E{row}+(E{row}*0.07)"
        assert worksheet.cell(row, mapping.columns["minimum"]).value == f"=E{row}-(E{row}*0.07)"
    assert "Control OCR" in workbook.sheetnames
    assert "Historial Obrixapy" in workbook.sheetnames


def test_dynamic_export_preserves_all_physical_slots_and_rejects_unapproved_final() -> None:
    mapping = analyse_template(synthetic_template())
    group = {
        "label": "T321",
        "label_number": 321,
        "classification": "distributed",
        "length_m": Decimal("31.496"),
        "strand_count": 8,
        "calculated_elongation": Decimal("20.6"),
        "measurements": [
            {
                "ordinal": ordinal,
                "measured_elongation": Decimal("20.6"),
                "review_status": "approved" if ordinal < 8 else "pending",
                "tolerance_status": "within",
            }
            for ordinal in range(1, 9)
        ],
    }
    with pytest.raises(TemplateValidationError, match="sin aprobación"):
        build_export_xlsx(
            synthetic_template(), mapping, [group], final=True, history={"kind": "final"}
        )
    group["measurements"][-1]["review_status"] = "approved"
    content = build_export_xlsx(
        synthetic_template(), mapping, [group], final=True, history={"kind": "final"}
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    rows = [
        row
        for row in range(1, workbook["Operativa"].max_row + 1)
        if str(workbook["Operativa"].cell(row, mapping.columns["calculated"]).value) == "20.6"
    ]
    assert len(rows) == 8
    assert len(set(rows)) == 8


def test_template_rejects_macros_and_external_like_payloads() -> None:
    malicious = BytesIO()
    with ZipFile(malicious, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types />")
        archive.writestr("xl/workbook.xml", "<workbook />")
        archive.writestr("xl/vbaProject.bin", b"macro")
    with pytest.raises(TemplateValidationError, match="macros"):
        validate_template_bytes(malicious.getvalue(), "macro.xlsx")


def test_template_rejects_external_links_before_openpyxl_reads_them() -> None:
    malicious = BytesIO()
    with ZipFile(malicious, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types />")
        archive.writestr("xl/workbook.xml", "<workbook />")
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            (
                '<Relationships><Relationship TargetMode="External" '
                'Target="https://example.test" /></Relationships>'
            ),
        )
    with pytest.raises(TemplateValidationError, match="enlaces externos"):
        validate_template_bytes(malicious.getvalue(), "externo.xlsx")


def test_invalid_formula_reference_is_not_a_valid_template_seed() -> None:
    workbook = load_workbook(BytesIO(synthetic_template()), data_only=False)
    worksheet = workbook["Operativa"]
    worksheet["F5"] = "=E4+(E4*0.07)"
    worksheet["H5"] = "=E4-(E4*0.07)"
    content = BytesIO()
    workbook.save(content)
    with pytest.raises(TemplateValidationError, match="fórmulas"):
        analyse_template(content.getvalue(), "referencia.xlsx")


def test_progress_blocks_outside_measurement_without_observation() -> None:
    job = SimpleNamespace(tolerance_percent=Decimal("7"), theory_approved_at=object())
    item = SimpleNamespace(
        id="item-1",
        strand_count=1,
        classification="distributed",
        theory_review_status="approved",
        calculated_elongation=Decimal("10"),
    )
    measurement = SimpleNamespace(
        item_id="item-1",
        measured_elongation=Decimal("10.8"),
        review_status="approved",
        override_reason=None,
    )
    progress = progress_for(job, [item], [measurement], [])
    assert progress["outside_tolerance"] == 1
    assert progress["can_approve_final"] is False
    assert "justificación" in " ".join(progress["approval_blockers"])


def test_security_rejects_an_xlsx_mime_with_non_zip_content() -> None:
    upload = UploadFile(
        filename="falso.xlsx",
        file=BytesIO(b"esto no es un zip"),
        headers=Headers({"content-type": next(iter(XLSX_MIME_TYPES))}),
    )
    with pytest.raises(HTTPException, match="no coincide"):
        asyncio.run(
            store_upload(upload, "companies", "company-1", allowed_mime_types=XLSX_MIME_TYPES)
        )


def test_job_lookup_is_scoped_to_its_company_and_project() -> None:
    statements: list[str] = []

    class ScopedSession:
        async def scalar(self, statement: object) -> None:
            statements.append(str(statement))
            return None

    with pytest.raises(HTTPException, match="no encontrado") as error:
        asyncio.run(require_job(ScopedSession(), "company-a", "project-a", "job-b"))
    assert error.value.status_code == 404
    assert "elongation_jobs.company_id" in statements[0]
    assert "elongation_jobs.project_id" in statements[0]
    assert "supervisor" not in DOCUMENT_APPROVER_ROLES


def test_json_snapshots_keep_decimal_precision_as_strings() -> None:
    assert json_safe({"calculated": Decimal("7.900")}) == {"calculated": "7.900"}
